from datetime import datetime
from html import escape
import os
import re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


BRAND_NAVY = "0B1220"
BRAND_PANEL = "132238"
BRAND_BLUE = "3B82F6"
BRAND_GREEN = "10B981"
BRAND_RED = "EF4444"
SOFT_BLUE = "EAF2FF"
SOFT_ROW = "F7FAFC"
BORDER = "D7DEE8"


def clean_text_for_pdf(text):
    if not isinstance(text, str):
        return str(text) if text is not None else ""
    return re.sub(r"[^\u0000-\u007F\u00C0-\u00FF\u0100-\u017F\u20AC]", "", text).strip()


def _normal_rows(rows):
    return [list(row) for row in rows]


def _metadata_items(metadata, rows):
    items = [("Oluşturma tarihi", datetime.now().strftime("%d/%m/%Y %H:%M")), ("Kayıt sayısı", str(len(rows)))]
    for item in metadata or []:
        if len(item) >= 2:
            items.append((str(item[0]), str(item[1])))
    return items


def _safe_sheet_name(name):
    safe = re.sub(r"[\[\]\:\*\?\/\\]", " ", str(name or "Rapor")).strip() or "Rapor"
    return safe[:31]


def _safe_table_name(title):
    safe = re.sub(r"[^A-Za-z0-9_]", "_", str(title or "Rapor"))
    safe = re.sub(r"_+", "_", safe).strip("_") or "Rapor"
    if safe[0].isdigit():
        safe = f"T_{safe}"
    return safe[:240]


def _estimate_excel_width(values, column_count):
    max_len = 8
    for value in values:
        max_len = max(max_len, len(str(value or "")))
    cap = 42 if column_count <= 5 else 32 if column_count <= 8 else 24
    return min(max(max_len + 3, 12), cap)


def export_rows_to_excel(file_path, title, columns, rows, subtitle=None, metadata=None, sheet_name="Rapor"):
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("Excel çıktısı için openpyxl kurulu olmalıdır.")

    rows = _normal_rows(rows)
    columns = list(columns)
    column_count = max(1, len(columns))
    last_col = get_column_letter(column_count)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _safe_sheet_name(sheet_name)
    sheet.sheet_view.showGridLines = False

    title_text = str(title)
    subtitle_text = subtitle or "ALP Ziraat Hayvan Yönetim Platformu"

    sheet.append([title_text])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    sheet.row_dimensions[1].height = 30
    sheet["A1"].font = Font(bold=True, size=17, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor=BRAND_NAVY)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")

    sheet.append([subtitle_text])
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=column_count)
    sheet.row_dimensions[2].height = 24
    sheet["A2"].font = Font(size=10, color="D9E8FF")
    sheet["A2"].fill = PatternFill("solid", fgColor=BRAND_PANEL)
    sheet["A2"].alignment = Alignment(horizontal="left", vertical="center")

    for row in (1, 2):
        for col in range(1, column_count + 1):
            cell = sheet.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=BRAND_NAVY if row == 1 else BRAND_PANEL)

    meta_fill = PatternFill("solid", fgColor=SOFT_BLUE)
    value_fill = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color=BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for label, value in _metadata_items(metadata, rows):
        row_idx = sheet.max_row + 1
        sheet.append([label, value])
        sheet.cell(row=row_idx, column=1).font = Font(bold=True, color=BRAND_PANEL)
        sheet.cell(row=row_idx, column=1).fill = meta_fill
        sheet.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")
        sheet.cell(row=row_idx, column=2).fill = value_fill
        sheet.cell(row=row_idx, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if column_count > 2:
            sheet.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=column_count)
        for col in range(1, column_count + 1):
            sheet.cell(row=row_idx, column=col).border = border

    sheet.append([""] * column_count)
    sheet.append([str(column) for column in columns])
    header_row = sheet.max_row

    header_fill = PatternFill("solid", fgColor=BRAND_PANEL)
    for cell in sheet[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    sheet.row_dimensions[header_row].height = 24

    for row in rows:
        safe_row = list(row[:column_count]) + [""] * max(0, column_count - len(row))
        sheet.append(safe_row)

    data_start = header_row + 1
    data_end = sheet.max_row
    for row_idx in range(data_start, data_end + 1):
        fill = PatternFill("solid", fgColor="FFFFFF" if (row_idx - data_start) % 2 == 0 else SOFT_ROW)
        for col_idx in range(1, column_count + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    sample_rows = rows[:200]
    for column_index, column_title in enumerate(columns, 1):
        values = [column_title] + [row[column_index - 1] if len(row) >= column_index else "" for row in sample_rows]
        sheet.column_dimensions[get_column_letter(column_index)].width = _estimate_excel_width(values, column_count)

    sheet.freeze_panes = f"A{data_start}"
    sheet.auto_filter.ref = f"A{header_row}:{last_col}{max(data_end, header_row)}"
    sheet.print_title_rows = f"{header_row}:{header_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.45
    sheet.page_margins.bottom = 0.45

    workbook.save(str(file_path))


def _register_pdf_fonts(styles):
    font_name = "Helvetica"
    font_bold = "Helvetica-Bold"
    arial_path = os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "arial.ttf")
    arial_bd_path = os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "arialbd.ttf")

    if os.path.exists(arial_path) and os.path.exists(arial_bd_path):
        try:
            pdfmetrics.registerFont(TTFont("Arial_TR", arial_path))
            pdfmetrics.registerFont(TTFont("Arial_TR-Bold", arial_bd_path))
            font_name = "Arial_TR"
            font_bold = "Arial_TR-Bold"
            styles["Normal"].fontName = font_name
        except Exception:
            pass
    return font_name, font_bold


def _paragraph(text, style):
    clean = escape(clean_text_for_pdf(text)).replace("\n", "<br/>")
    return Paragraph(clean or "-", style)


def _pdf_col_widths(columns, rows, available_width):
    lengths = []
    for idx, column in enumerate(columns):
        values = [column] + [row[idx] if len(row) > idx else "" for row in rows[:100]]
        longest = max(len(str(value or "")) for value in values)
        lengths.append(max(8, min(longest, 34)))
    total = sum(lengths) or 1
    widths = [available_width * (length / total) for length in lengths]
    min_width = 46 if len(columns) > 8 else 62
    if any(width < min_width for width in widths):
        widths = [max(width, min_width) for width in widths]
        scale = available_width / sum(widths)
        widths = [width * scale for width in widths]
    return widths


def export_rows_to_pdf(file_path, title, columns, rows, subtitle=None, metadata=None, sheet_name="Rapor"):
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("PDF çıktısı için reportlab kurulu olmalıdır.")

    rows = _normal_rows(rows)
    columns = list(columns)

    doc = SimpleDocTemplate(
        str(file_path),
        pagesize=landscape(A4),
        rightMargin=20,
        leftMargin=20,
        topMargin=22,
        bottomMargin=22,
        title=clean_text_for_pdf(title),
    )
    available_width = doc.pagesize[0] - doc.leftMargin - doc.rightMargin
    styles = getSampleStyleSheet()
    font_name, font_bold = _register_pdf_fonts(styles)

    title_style = ParagraphStyle(
        "AlpTitle",
        parent=styles["Normal"],
        fontName=font_bold,
        fontSize=15,
        leading=18,
        textColor=colors.white,
        alignment=TA_LEFT,
    )
    brand_style = ParagraphStyle(
        "AlpBrand",
        parent=styles["Normal"],
        fontName=font_bold,
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#D9E8FF"),
        alignment=TA_CENTER,
    )
    small_style = ParagraphStyle(
        "AlpSmall",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#334155"),
    )
    header_style = ParagraphStyle(
        "AlpTableHeader",
        parent=styles["Normal"],
        fontName=font_bold,
        fontSize=7.5 if len(columns) > 8 else 8.5,
        leading=9,
        textColor=colors.white,
        alignment=TA_CENTER,
    )
    cell_style = ParagraphStyle(
        "AlpTableCell",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=7 if len(columns) > 8 else 8,
        leading=8.5 if len(columns) > 8 else 9.5,
        textColor=colors.HexColor("#0F172A"),
        alignment=TA_LEFT,
    )

    story = []
    title_table = Table(
        [[_paragraph("ALP Ziraat", brand_style), _paragraph(title, title_style)]],
        colWidths=[95, available_width - 95],
    )
    title_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(f"#{BRAND_NAVY}")),
        ("LINEBELOW", (0, 0), (-1, -1), 2, colors.HexColor(f"#{BRAND_GREEN}")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(title_table)
    story.append(Spacer(1, 8))

    subtitle_text = subtitle or "ALP Ziraat Hayvan Yönetim Platformu"
    story.append(Paragraph(escape(clean_text_for_pdf(subtitle_text)), small_style))
    story.append(Spacer(1, 8))

    meta_rows = []
    items = _metadata_items(metadata, rows)
    for idx in range(0, len(items), 2):
        first = items[idx]
        second = items[idx + 1] if idx + 1 < len(items) else ("", "")
        meta_rows.append([
            _paragraph(first[0], small_style),
            _paragraph(first[1], small_style),
            _paragraph(second[0], small_style) if second[0] else "",
            _paragraph(second[1], small_style) if second[1] else "",
        ])
    meta_table = Table(meta_rows, colWidths=[90, (available_width - 180) / 2, 90, (available_width - 180) / 2])
    meta_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor(f"#{SOFT_BLUE}")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor(f"#{SOFT_BLUE}")),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor(f"#{BORDER}")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 10))

    table_data = [[_paragraph(column, header_style) for column in columns]]
    for row in rows:
        safe_row = list(row[:len(columns)]) + [""] * max(0, len(columns) - len(row))
        table_data.append([_paragraph(cell, cell_style) for cell in safe_row])

    table = Table(table_data, colWidths=_pdf_col_widths(columns, rows, available_width), repeatRows=1, splitByRow=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{BRAND_PANEL}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor(f"#{BORDER}")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(f"#{SOFT_ROW}")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(table)

    def footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont(font_name, 7)
        canvas.setFillColor(colors.HexColor("#64748B"))
        canvas.drawString(doc_obj.leftMargin, 10, clean_text_for_pdf(sheet_name))
        canvas.drawRightString(doc_obj.pagesize[0] - doc_obj.rightMargin, 10, f"Sayfa {doc_obj.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
