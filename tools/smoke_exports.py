import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]


def main():
    import sys
    sys.path.insert(0, str(ROOT))

    from alp_ziraat_export import export_rows_to_excel, export_rows_to_pdf

    tmp = Path(tempfile.mkdtemp(prefix="alp_export_smoke_"))
    xlsx_path = tmp / "hayvan_listesi.xlsx"
    pdf_path = tmp / "hayvan_listesi.pdf"

    columns = ["Çiftlik", "Resmi Küpe", "Çiftlik Küpesi", "Irk", "Yaş", "Cinsi", "Durum", "Uyarılar"]
    rows = [
        ["Sametin Çiftliği", "TR001", "C001", "Simental", "2 yıl 1 ay", "Düve", "Gebe", "Gebelik kontrolü yaklaşıyor"],
        ["Sametin Çiftliği", "TR002", "C002", "Holstein", "8 ay", "Dana", "Aktif", "-"],
    ]
    metadata = [("Kullanıcı", "admin"), ("Çalışılan alan", "Sametin Çiftliği"), ("Bağlantı", "Online")]

    export_rows_to_excel(
        xlsx_path,
        "ALP Ziraat Hayvan Listesi",
        columns,
        rows,
        subtitle="Hayvan listesindeki mevcut görünüm",
        metadata=metadata,
        sheet_name="Hayvan Listesi",
    )
    export_rows_to_pdf(
        pdf_path,
        "ALP Ziraat Hayvan Listesi",
        columns,
        rows,
        subtitle="Hayvan listesindeki mevcut görünüm",
        metadata=metadata,
        sheet_name="Hayvan Listesi",
    )

    assert xlsx_path.exists() and xlsx_path.stat().st_size > 5000
    assert pdf_path.exists() and pdf_path.stat().st_size > 2500

    workbook = load_workbook(xlsx_path)
    sheet = workbook["Hayvan Listesi"]
    assert sheet["A1"].value == "ALP Ziraat Hayvan Listesi"
    assert sheet.freeze_panes
    assert any(sheet.cell(row=9, column=col).value == "Irk" for col in range(1, sheet.max_column + 1))
    assert "Kayıt sayısı" in [sheet.cell(row=i, column=1).value for i in range(1, 8)]
    assert sheet.max_row >= 8
    with zipfile.ZipFile(xlsx_path) as archive:
        names = archive.namelist()
        assert not any(name.startswith("xl/tables/") for name in names), names
        worksheet_xml = [
            archive.read(name).decode("utf-8", errors="ignore")
            for name in names
            if name.startswith("xl/worksheets/")
        ]
        assert not any("tableParts" in xml for xml in worksheet_xml)

    print(f"Export smoke OK: {tmp}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
